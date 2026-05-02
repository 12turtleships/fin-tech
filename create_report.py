import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime

# Read the existing Excel file with the correct sheet name (lowercase)
file_path = "/path/to/4D_Leadership_List.xlsx"
df = pd.read_excel(file_path, sheet_name="Group Leaders")

# Convert appointment date to datetime
df['Date Appointed'] = pd.to_datetime(df['Date Appointed'])

# Filter for year 2025
df_2025 = df[df['Date Appointed'].dt.year == 2025]

# Create summary by region, chapter and month
summary = pd.pivot_table(
    df_2025,
    index=['Region', 'Chapter'],
    columns=df_2025['Date Appointed'].dt.strftime('%b'),
    aggfunc='size',
    fill_value=0
).reset_index()

# Ensure Jan, Feb, Mar columns exist
for month in ['Jan', 'Feb', 'Mar']:
    if month not in summary.columns:
        summary[month] = 0

# Calculate Total YTD
summary['Total YTD'] = summary[['Jan', 'Feb', 'Mar']].sum(axis=1)

# Reorder columns
summary = summary[['Region', 'Chapter', 'Jan', 'Feb', 'Mar', 'Total YTD']]

# Calculate region subtotals
region_totals = summary.groupby('Region').agg({
    'Jan': 'sum',
    'Feb': 'sum',
    'Mar': 'sum',
    'Total YTD': 'sum'
}).reset_index()

# Create Mid-Atlantic total row
total_row = pd.DataFrame({
    'Region': ['Mid-Atlantic'],
    'Chapter': ['Total'],
    'Jan': [summary['Jan'].sum()],
    'Feb': [summary['Feb'].sum()],
    'Mar': [summary['Mar'].sum()],
    'Total YTD': [summary['Total YTD'].sum()]
})

# Add region subtotals
final_summary = []
for region in region_totals['Region'].unique():
    # Add chapter rows for this region
    region_chapters = summary[summary['Region'] == region].copy()
    final_summary.append(region_chapters)
    
    # Add region subtotal
    region_total = region_totals[region_totals['Region'] == region].copy()
    region_total['Chapter'] = 'Subtotal'
    final_summary.append(region_total)

# Combine all rows and add Mid-Atlantic total
final_summary = pd.concat(final_summary + [total_row], ignore_index=True)

# Create Excel writer object
with pd.ExcelWriter('group_leaders_report.xlsx', engine='openpyxl') as writer:
    # Write DataFrame to Excel
    final_summary.to_excel(writer, sheet_name='Group Leaders Report', index=False)
    
    # Get the workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['Group Leaders Report']
    
    # Format headers
    for col in range(len(final_summary.columns)):
        cell = worksheet.cell(row=1, column=col+1)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Format subtotal and total rows
    for row in range(2, len(final_summary) + 2):  # +2 because Excel is 1-indexed and we have a header row
        if worksheet.cell(row=row, column=2).value in ['Subtotal', 'Total']:
            for col in range(1, len(final_summary.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

print("Excel file 'group_leaders_report.xlsx' has been created successfully.") 